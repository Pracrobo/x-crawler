import os
from dotenv import load_dotenv
from logger import get_logger
from openpyxl import load_workbook
from playwright.async_api import async_playwright, TimeoutError
import asyncio
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
import time 

logger = get_logger()
load_dotenv()
CURRENT_TIME = datetime.now()

def urlCollecter():
    wb= load_workbook(os.getenv("TARGET_FILE_PATH"))
    ws = wb.active

    col_a, col_c = 'A', 'C'
    urls_group = []
    group = []
    empty_row_count = 0
    row = 3

    while True:
        a_value = ws[f'{col_a}{row}'].value
        c_value = ws[f'{col_c}{row}'].value
        if c_value is None:
            if group:
                urls_group.append(group)
                group = []
                time.sleep(1)
            empty_row_count += 1
            if empty_row_count >= 3:
                break
        else:
            empty_row_count = 0
            pair = (a_value, c_value)
            group.append(pair)

            if len(group) == 5:     
                urls_group.append(group)
                group = []
                time.sleep(1)
        row += 1
    
    if group:
        urls_group.append(group)
    total_items = sum(len(g) for g in urls_group)
    logger.info(f"총 {len(urls_group)} 단위, {total_items}개 url 모으기 끝")

    return urls_group


async def extract_elements(content):
    result = {
        "id" : None,
        "time": None,
        "mention": None,
        "views_info": []
    }

    datetime_str = None
    clean_string = []
    words = [e.strip() for e in content]
    for word in words:
        # id 찾기
        if word.startswith("@"):
            result["id"] = word
        # 날짜 변환
        if "AM" in word or "PM" in word:
            if "·" in word and "," in word:
                datetime_str = word.strip()
                # datetime 객체로 변환
                if datetime_str:
                    clean_dt = " ".join([x.strip() for x in datetime_str.split("·")])
                    dt = datetime.strptime(clean_dt, "%I:%M %p %b %d, %Y")
                    result["time"] = dt.strftime("%Y-%m-%d %H:%M:%S")

        if "Views" in word:
            idx = words.index("Views")
            filtered = words[idx-1:idx+3]
            filtered.remove("Views")
            result["views_info"] = filtered
        
        if "Translate post" in word:
            idx = words.index("Translate post")
            clean_string = words[2:idx]
            result["mention"] = '\n'.join([e.strip() for e in clean_string if e.strip()])

    return result

def failed_urls(url, error,  filename=(os.getenv("FILE_PATH_2"))):
    """에러난 URL을 지정된 텍스트 파일에 한 줄씩 저장"""
    file_dir = os.getenv("FAIL_PATH")
    file_path = os.path.join(file_dir, f"{CURRENT_TIME.date()}-{filename}")
    try:
        with open(file_path, "a", encoding="utf-8") as f:
            f.write(f"{url}: {error}\n")
    except Exception as e:
        logger.error(f"[URL 저장 실패] {url} → {e}")

def is_valid_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme in ('http', 'https'), result.netloc])
    except:
        return False
    

def find_file(filename=(os.getenv("FILE_PATH_2"))):
    """멈춘 파일이 있는지 확인"""
    urls = []
    file_dir=os.getenv("FAIL_PATH")
    file_path = os.path.join(file_dir, f"{CURRENT_TIME.date()}-{filename}")
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            if ':' in line:
                parts = line.strip().split(':', 2)  # 최대 2번만 split
                if len(parts) >= 2:
                    # 앞의 두 파트는 URL일 수 있으므로 다시 합침
                    maybe_url = ':'.join(parts[:2]).strip()
                    if is_valid_url(maybe_url):
                        urls.append(maybe_url)

    chunked_urls = chunk_list(urls, chunk_size=5)
    return chunked_urls


def chunk_list(lst, chunk_size=5):
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]
    

def to_excel(data_list):
    """결과 excel로 저장하기"""

    file_dir = os.getenv("FILE_PATH_1")
    os.makedirs(file_dir, exist_ok=True)  # 폴더 없으면 자동 생성

    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(file_dir, f"{date_str}.xlsx")

    rows = []
    for data in data_list:
        views_values = (data.get("views_info") or []) + [None, None, None]
        views_values = views_values[:3]
        row = {
        "No": data.get("no"),
        "멘션" : data.get("mention"),
        "게재 URL" : data.get("url"),
        "계정": data.get("id"),
        "게시날짜": data.get("time"),
        "조회수": views_values[0],
        "재게시": views_values[1],
        "좋아요": views_values[2]
        } 
        rows.append(row)

    df = pd.DataFrame(rows)
    
    try:
        if not os.path.exists(filename):
            df.to_excel(filename, index=False)
        else:
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as w:
                sheet = w.sheets["Sheet1"]
                df.to_excel(w, index=False, header=False, startrow=sheet.max_row)
                logger.info("행 단위 청크 엑셀 저장 완료")

    except Exception as e:
        logger.error(f"에러: {e}", exc_info=True)
        if not df.empty:
            url = df.iloc[0].get("게재 URL", "알 수 없음")
            failed_urls(url, e)
            logger.error(f"에러가 발생한 URL: {url}")


async def process_page(page, url):
    logger.debug(f"접속 시도: {url}")
    await page.goto(url, timeout=30000) # Option DOM요소로딩 후 시작하게끔
    locator = page.locator("article")
    await locator.wait_for(state="visible", timeout=60000)
    content = await page.inner_text("article")
    result = await extract_elements(content.split("\n"))  
    result['url'] = url
    return result


async def retry_url(page, no, url, retries=3):
    for attempt in range(1, retries + 1):
        try:
            result = await process_page(page, url)
            result["no"] = no
            return {"result": result, "url": url, "error": None}
        except Exception as e:
            logger.warning(f"{url} 처리 실패 {attempt}/{retries}: {e}")
            if attempt == retries:
                logger.error(f"{url} 최종 실패")
                failed_urls(url, e)
                return {"result": None, "url": url, "error": e}
            await asyncio.sleep(10)  # 재시도 전 잠깐 대기


async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        file_path  = os.getenv("FILE_PATH_2")
        file_dir = os.getenv("FAIL_PATH")  # 경로 변경
        os.makedirs(file_dir, exist_ok=True) 
        filename = os.path.join(file_dir, f"{CURRENT_TIME.date()}-{file_path}")
        if os.path.exists(filename):
            logger.info("실패한 파일 목록이 존재합니다.")
            urls_group = find_file()
        else:
            urls_group = urlCollecter() 
            
        
        for group in urls_group:
            tasks = []
            pages = []
            logger.info('group 단위 페이지 탭 열기 시작')

            for no, url in group:
                page = await context.new_page()
                pages.append(page)
                tasks.append(retry_url(page, no, url))

            results = await asyncio.gather(*tasks)

            success_results = [r["result"] for r in results if r["error"] is None]
            error_results = [r for r in results if r["error"] is not None]

            if success_results:
                to_excel(success_results)

            for err in error_results:
                logger.error(f"처리 실패 URL: {err['url']} 에러: {err['error']}")

            for page in pages:
                await page.close()

        logger.info('브라우저 닫기')
        await browser.close()


asyncio.run(main())
